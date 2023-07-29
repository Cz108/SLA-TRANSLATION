import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from tools.colors import print_colored_block
import sys

"""
 In Excel, rows and columns are 1-indexed, meaning they start from 1, not 0.
"""

def load_workbook(file_path):
    return openpyxl.load_workbook(file_path)

def get_cell_dict_in_selected_column(sheet, col_index=1) -> list:
    # sheet = workbook.worksheets[table_index]
    cell_dict_in_selected_column = []
    for row_index, row in enumerate(sheet.iter_rows(), start=1):
        cell = row[col_index]
        cell_dict_in_selected_column.append({'object': cell, 'index': (sheet.title, row_index, col_index)})
    return cell_dict_in_selected_column

def find_cell_by_index(sheet, index):
    row_index, col_index = index
    try:
        cell = sheet.cell(row=row_index, column=col_index)
        return {'object': cell, 'index': index}
    except IndexError:
        return None

def copy_font(source_font, target_font):
    target_font.name = source_font.name
    target_font.size = source_font.size
    target_font.bold = source_font.bold
    target_font.italic = source_font.italic
    target_font.underline = source_font.underline
    target_font.color = source_font.color

def copy_cell_format(source_cell, target_cell):
    copy_font(source_cell.font, target_cell.font)

def edit_cell_content(sheet, index, new_content):
    table_index, row_index, col_index = index
    try:
        cell = sheet.cell(row=row_index, column=col_index + 1)
        new_cell = sheet.cell(row=row_index, column=col_index + 1)
        # copy_cell_format(cell, new_cell)
        new_cell.value = new_content
        return True
    except IndexError:
        return False

def delete_cell_content(sheet, index):
    row_index, col_index = index
    try:
        cell = sheet.cell(row=row_index, column=col_index)
        cell.value = None
        return True
    except IndexError:
        return False


def get_cell_shading_color(cell):
    tmp = cell.fill.fill_type
    if cell.fill.fill_type != "none":
        argb_color = cell.fill.start_color.index
        rgb_color = argb_color[2:]
        # Return the actual color, otherwise return white color
        return rgb_color if rgb_color != "000000" else "FFFFFF"
    # White color, no color, return "FFFFFF"
    return "FFFFFF"


def check_table_colors(sheet):
    table_colors = set()
    for row in sheet.iter_rows():
        for cell in row:
            color = get_cell_shading_color(cell)
            if color:
                table_colors.add(color)

    for color in table_colors:
        print_colored_block(color)

    return table_colors


def get_cell_text(cell):
    return cell.value


def print_cell_color(cell):
    color = get_cell_shading_color(cell)
    if color:
        print_colored_block(color)


def if_color_match_the_cell(cell, color_to_be_checked):
    cell_color = get_cell_shading_color(cell)
    return True if cell_color == color_to_be_checked else False


def column_processor(sheet, column_index_to_select=0):
    cells_in_selected_column = get_cell_dict_in_selected_column(sheet, column_index_to_select)
    for cell_dict in cells_in_selected_column:
        print("-" * 50)
        print(get_cell_text(cell_dict['object']))
        print_cell_color(cell_dict['object'])


if __name__ == "__main__":
    sys.stdout.reconfigure(encoding='utf-8')
    excel_path = '/Users/bilibala/Study/SLA-TRANS/documents/excel_test.xlsx'
    wb = load_workbook(excel_path)
    sheet = wb.active

    print("\nINFO: column_processor(sheet, 0)")
    column_processor(sheet, 0)

    print("\nINFO: check_table_colors(sheet)")
    table_colors = check_table_colors(sheet)

    print("\nINFO: edit_cell_content")
    content = "test abc content"
    if edit_cell_content(sheet, (0, 2, 1), content):
        print("Cell content modified successfully.")
    else:
        print("Failed to modify cell content.")


    wb.save(excel_path)
